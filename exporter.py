import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)


# ── Colour palette (hex, no #) ────────────────────────────────────────────────
C_HEADER_BG   = "0F3460"
C_HEADER_FG   = "00D4FF"
C_SECTION_BG  = "16213E"
C_SECTION_FG  = "00D4FF"
C_ROW_BG_ODD  = "1A1A2E"
C_ROW_BG_EVEN = "16213E"
C_FG_DEFAULT  = "EEEEEE"
C_FG_GREEN    = "00FF88"
C_FG_AMBER    = "FFAA00"
C_FG_RED      = "FF4757"

FLAG_COLOUR_MAP = {
    "standard": C_FG_GREEN,
    "notable":  C_FG_AMBER,
    "unknown":  C_FG_RED,
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _font(hex_colour=C_FG_DEFAULT, bold=False, size=10):
    return Font(color=hex_colour, bold=bold, size=size, name="Calibri")

def _border():
    side = Side(style="thin", color="0F3460")
    return Border(left=side, right=side, top=side, bottom=side)

def _apply_header_row(ws, row_num, values, col_widths=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _font(C_HEADER_FG, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[row_num].height = 20

def _apply_data_row(ws, row_num, values, colours=None):
    bg = C_ROW_BG_ODD if row_num % 2 == 1 else C_ROW_BG_EVEN
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = _fill(bg)
        fg = colours[col - 1] if colours and col - 1 < len(colours) else C_FG_DEFAULT
        cell.font      = _font(fg)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[row_num].height = 18

def _autofit_columns(ws, min_width=10, max_width=60, padding=2):
    """
    Iterate every column in the sheet and set its width to the longest
    cell value found, clamped between min_width and max_width.
    padding adds a small breathing gap so text isn't flush with the border.
    """
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                # Handle multi-line values — use the longest line
                lines = str(cell.value).split("\n")
                cell_len = max(len(line) for line in lines)
                if cell_len > max_len:
                    max_len = cell_len
        width = min(max(max_len + padding, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _tab_colour(ws, hex_colour):
    ws.sheet_properties.tabColor = hex_colour


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_port_sheet(wb, port_scans):
    ws = wb.active
    ws.title = "Port Scans"
    ws.sheet_view.showGridLines = False
    _tab_colour(ws, C_HEADER_FG)
    ws.sheet_properties.tabColor = "00D4FF"

    headers = ["Timestamp", "Protocol", "Switch", "Port",
               "Model / Description", "Management IP", "VLAN"]

    _apply_header_row(ws, 1, headers)
    _freeze(ws)

    for i, entry in enumerate(port_scans, start=2):
        proto_colour = C_HEADER_FG if entry["protocol"] == "LLDP" else "FF9500"
        values = [
            entry["timestamp"],
            entry["protocol"],
            entry["switch"],
            entry["port"],
            entry["model"],
            entry["ip"],
            entry["vlan"],
        ]
        colours = [C_FG_DEFAULT, proto_colour] + [C_FG_DEFAULT] * 5
        _apply_data_row(ws, i, values, colours)

    # Summary row at the bottom
    if port_scans:
        summary_row = len(port_scans) + 3
        cell = ws.cell(row=summary_row, column=1,
                       value=f"Total scans: {len(port_scans)}")
        cell.font = _font(C_HEADER_FG, bold=True)
        cell.fill = _fill(C_SECTION_BG)

    _autofit_columns(ws)


def _build_ip_sheet(wb, ip_snapshots):
    ws = wb.create_sheet("IP Snapshots")
    ws.sheet_view.showGridLines = False
    _tab_colour(ws, "00FF88")

    summary_headers = ["Timestamp", "Adapter", "Hostname", "MAC",
                        "IP Address", "Subnet", "Gateway",
                        "DNS Servers", "Domain", "DHCP Enabled",
                        "DHCP Server", "Lease Obtained", "Lease Expires"]

    _apply_header_row(ws, 1, summary_headers)
    _freeze(ws)

    for i, snap in enumerate(ip_snapshots, start=2):
        values = [
            snap["timestamp"],
            snap["adapter"],
            snap["hostname"],
            snap["mac"],
            snap["ip"],
            snap["subnet"],
            snap["gateway"],
            snap["dns"],
            snap["domain"],
            "Yes" if snap["dhcp_enabled"] else "No",
            snap["dhcp_server"],
            snap["lease_obtained"],
            snap["lease_expires"],
        ]
        dhcp_col = C_FG_GREEN if snap["dhcp_enabled"] else C_FG_RED
        colours = [C_FG_DEFAULT] * 9 + [dhcp_col] + [C_FG_DEFAULT] * 3
        _apply_data_row(ws, i, values, colours)

    # ── DHCP options section below ──
    if not ip_snapshots:
        _autofit_columns(ws)
        return

    current_row = len(ip_snapshots) + 4

    # Section header
    cell = ws.cell(row=current_row, column=1, value="DHCP Scope Options")
    cell.font = _font(C_HEADER_FG, bold=True, size=11)
    cell.fill = _fill(C_SECTION_BG)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    opt_headers = ["Snapshot #", "Option #", "Option Name", "Value", "Flag"]
    _apply_header_row(ws, current_row, opt_headers)
    current_row += 1

    for snap_num, snap in enumerate(ip_snapshots, start=1):
        for opt_num, label, value, flag in snap.get("dhcp_options", []):
            fg = FLAG_COLOUR_MAP.get(flag, C_FG_DEFAULT)
            values  = [f"Snapshot {snap_num}", opt_num, label, value, flag.capitalize()]
            colours = [C_FG_DEFAULT, C_FG_DEFAULT, C_FG_DEFAULT, fg, fg]
            _apply_data_row(ws, current_row, values, colours)
            current_row += 1

    _autofit_columns(ws)


def _build_mdns_sheet(wb, mdns_scans):
    ws = wb.create_sheet("mDNS Discoveries")
    ws.sheet_view.showGridLines = False
    _tab_colour(ws, "FF9500")

    headers = ["Scan #", "Timestamp", "Device Name", "Service Type",
               "IP Address", "Raw String"]

    _apply_header_row(ws, 1, headers)
    _freeze(ws)

    row = 2
    for scan_num, scan in enumerate(mdns_scans, start=1):
        for device in scan.get("devices", []):
            values = [
                scan_num,
                scan["timestamp"],
                device.get("friendly", ""),
                device.get("type", ""),
                device.get("ip", ""),
                device.get("raw", ""),
            ]
            _apply_data_row(ws, row, values)
            row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No mDNS data captured this session.").font = \
            _font(C_FG_AMBER)

    _autofit_columns(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def export_xlsx(session, filepath):
    """
    Export the full session to an .xlsx file.
    Always creates all three sheets regardless of what data is present.
    """
    wb = Workbook()

    _build_port_sheet(wb, session.port_scans)
    _build_ip_sheet(wb, session.ip_snapshots)
    _build_mdns_sheet(wb, session.mdns_scans)

    # Cover metadata in the Port Scans sheet
    ws = wb["Port Scans"]
    meta_row = 1
    ws.cell(row=meta_row, column=9,
            value=f"PiNT Export — {datetime.now().strftime('%d %b %Y %H:%M')}").font = \
        _font(C_HEADER_FG, bold=True)

    wb.save(filepath)


def export_csv(session, filepath):
    """
    Export Port Scans only to a flat .csv file.
    """
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Timestamp", "Protocol", "Switch", "Port",
                          "Model / Description", "Management IP", "VLAN"])
        for entry in session.port_scans:
            writer.writerow([
                entry["timestamp"],
                entry["protocol"],
                entry["switch"],
                entry["port"],
                entry["model"],
                entry["ip"],
                entry["vlan"],
            ])